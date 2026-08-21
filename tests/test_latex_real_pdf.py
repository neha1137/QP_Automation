"""
tests/test_latex_real_pdf.py — LaTeX regression tests anchored to the REAL
mock-test PDF (templates/sample_input.pdf), not synthetic strings.

These exist because unit tests on hand-written strings proved the detector
logic in isolation but did NOT prove that real PDF-extracted text (with its
own quirks — en-dashes used as both subtraction and arbitrary separators,
bare-digit visual-grid questions with no operators at all, letter-coded
"pseudo-operators" like P/Q/R/S, OCR-ish glyph swaps) reaches
latex_processor.py in a form it can act on correctly, and that the result
is what actually lands in Excel. Investigation against this real PDF found
and fixed two real bugs (see latex_processor.py's _tokenize docstring and
the trailing-slash-fraction retraction in detect_math_segments) — these
tests pin the corrected behavior against the exact real content that
exposed them, so a future change can't silently reintroduce either one.
"""

from __future__ import annotations

import re

import openpyxl

import parser as parser_mod
from excel_writer import write_excel, TEMPLATE_PATH
from paper_segmenter import detect_paper_spans

LATEX_SPAN_RE = re.compile(r"\$(.+?)\$")
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]{3,}")
# LaTeX command names are the only legitimate 3+-letter alphabetic runs
# that should ever appear inside a $...$ span.
_SAFE_LATEX_WORDS = {"times", "div", "le", "ge", "neq", "pm", "sqrt", "frac"}


def _real_questions():
    """Real, fully-processed (parser + LaTeX layer) questions from the
    actual mock-test PDF — session-scoped extraction is reused via the
    ssc_doc_result fixture where possible, but this helper also needs the
    PRE-latex snapshot, so it re-parses directly."""
    from conftest import SSC_PDF
    from extractor import extract_document

    doc_result = extract_document(SSC_PDF)
    spans = detect_paper_spans(doc_result)
    result = parser_mod.parse_paper(doc_result, spans[0])
    return {q["number"]: q for q in result["questions"]}


def _real_questions_before_latex():
    """Same real PDF, same parse, but with the LaTeX post-processing hook
    disabled — the raw parser output, for BEFORE/AFTER comparison."""
    from conftest import SSC_PDF
    from extractor import extract_document

    doc_result = extract_document(SSC_PDF)
    spans = detect_paper_spans(doc_result)
    orig = parser_mod.process_question_fields
    parser_mod.process_question_fields = lambda q: q
    try:
        result = parser_mod.parse_paper(doc_result, spans[0])
    finally:
        parser_mod.process_question_fields = orig
    return {q["number"]: q for q in result["questions"]}


# -- 1. Real multiplication/division/inequality expressions convert -------

def test_real_multiplication_equation_converted():
    q = _real_questions()[12]
    assert r"$4 \times 3 + 1 = 13$" in q["explanation"]
    assert r"$382 \times 3 + 5 = 1151$" in q["explanation"]


def test_real_division_and_multiplication_in_stem_and_explanation():
    q = _real_questions()[29]
    assert r"$\div$" in q["stem"]
    assert r"$\times$" in q["stem"]
    assert r"$14 \div 7 \times 6 + 4$" in q["explanation"]


def test_real_not_equal_symbol_converted():
    q = _real_questions()[29]
    assert r"$13 \neq 11$" in q["explanation"]


def test_real_letter_equation_pattern_converted():
    """'A + 2 = C' style alphabet-series reasoning — a genuine equation
    using single-letter variables, found in the real explanation text."""
    q = _real_questions()[7]
    assert r"$A + 2 = C$" in q["explanation"]
    assert r"$L + 2 = N$" in q["explanation"]


def test_real_weight_relation_equation_converted():
    q = _real_questions()[43]
    assert r"$P = 2Q$" in q["explanation"]
    assert r"$R = 3T$" in q["explanation"]


# -- 2. Regression: an ordinary word must never end up inside $...$ -------

def test_real_word_number_never_swallowed_into_latex_span():
    """Q13's explanation reads '...number × (1st number – 1)...' — the
    real bug found during investigation wrapped the English word "number"
    itself into the LaTeX span ('$number \\times$'). Only the operator may
    convert; the word must stay outside any $...$ delimiter."""
    q = _real_questions()[13]
    for span in LATEX_SPAN_RE.finditer(q["explanation"]):
        assert "number" not in span.group(1)
    assert r"$\times$" in q["explanation"] or r"number $\times$" in q["explanation"]
    assert "number $\\times$" in q["explanation"]


def test_no_english_word_ever_appears_inside_a_latex_span_whole_document():
    """Whole-document invariant, not just a spot check: scans every $...$
    span produced across all 200 real questions' stem/options/explanation
    and asserts none contains a 3+-letter English word (only LaTeX command
    names like \\times/\\div are legitimate multi-letter content)."""
    questions = _real_questions()
    offenders = []
    for q in questions.values():
        fields = [("stem", q.get("stem")), ("explanation", q.get("explanation"))]
        fields += [(f"option_{L}", q["options"].get(L)) for L in "ABCD"]
        for field_name, text in fields:
            if not text:
                continue
            for m in LATEX_SPAN_RE.finditer(text):
                bad = [w for w in ENGLISH_WORD_RE.findall(m.group(1)) if w.lower() not in _SAFE_LATEX_WORDS]
                if bad:
                    offenders.append((q["number"], field_name, m.group(0), bad))
    assert offenders == []


# -- 3. Regression: a fraction must never be clipped mid-expression -------

def test_real_embedded_fraction_never_clipped_by_a_dollar_sign():
    """Q29's explanation contains '...19.5 ≠ 35/2' — the real bug found
    during investigation produced '$19.5 \\neq 35$/2' (a $ landing between
    the numerator and the slash). The fraction must stay either fully
    outside $...$ (current, conservative behavior) or fully inside it —
    never split by a closing $ right before the slash."""
    q = _real_questions()[29]
    assert "$/" not in q["explanation"]
    assert "35/2" in q["explanation"]  # left as plain text, not fabricated


# -- 4. Correctly left as normal text -------------------------------------

def test_real_percentage_range_option_left_as_text():
    q = _real_questions()[66]
    assert q["options"]["A"] == "24-25%"
    assert q["options"]["B"] == "29-30%"


def test_real_letter_coded_pseudo_operator_option_left_as_text():
    """Q29 Option C: '3 S 6 P 2 Q 3 R 6 = 35/2' — P/Q/R/S are the
    question's OWN letter-codes for +,-,÷,× (defined in the stem), not
    real math notation, and the field isn't a whole-field fraction either
    — must stay completely untouched."""
    q = _real_questions()[29]
    assert q["options"]["C"] == "3 S 6 P 2 Q 3 R 6 = 35/2"


# -- 5. Ambiguous visual/grid content — intentionally never fabricated ----

def test_real_bare_digit_grid_stem_never_fabricates_operators():
    """Q37's stem is a visual number-grid question: PyMuPDF's native text
    extraction yields the grid's bare digits one per line with NO operator
    characters at all (the ×/=/pattern only exists in the explanation's
    prose, not the stem) — reconstructing "7 × 2 × 3 = 42" from layout
    alone would be fabrication. Confirms the stem is completely untouched
    (byte-for-byte identical before/after) and never modified."""
    before = _real_questions_before_latex()[37]
    after = _real_questions()[37]
    assert before["stem"] == after["stem"]
    assert "$" not in after["stem"]


def test_real_image_question_math_looking_text_left_untouched():
    """Q45 is a genuine image/visual question (is_image reflects the
    parser's own pre-LaTeX classification) whose explanation text is a
    fragment of bare digits with no operators. Confirms: (a) is_image is
    unaffected by LaTeX processing, (b) the bare-digit text is not
    converted into fabricated math, and (c) this is exactly the case
    where the LaTeX layer and the image-detection pipeline must coexist
    without either interfering with the other."""
    q = _real_questions()[45]
    assert q["is_image"] is True
    assert "$" not in q["explanation"]


def test_image_question_set_unaffected_by_latex_processing():
    """The real document's is_image classification (computed in
    finalize_questions(), BEFORE the LaTeX hook runs) must be identical
    with or without LaTeX processing — confirms the two systems are
    architecturally independent for this real PDF, not just in theory."""
    before = _real_questions_before_latex()
    after = _real_questions()
    before_image_qs = {n for n, q in before.items() if q["is_image"]}
    after_image_qs = {n for n, q in after.items() if q["is_image"]}
    assert before_image_qs == after_image_qs == {5, 39, 44, 45, 46, 47}


# -- 6. Excel: real PDF -> parser -> LaTeX -> Excel cell, end to end ------

def test_real_pdf_end_to_end_excel_contains_latex_as_plain_text(tmp_path):
    """The full real pipeline: extract templates/sample_input.pdf, parse,
    run the (already-integrated) LaTeX hook, write Excel, then verify
    with openpyxl that the LaTeX landed in the right cells as plain TEXT
    (never an Excel formula, never a rendered image)."""
    from validator import validate_document

    questions = list(_real_questions().values())
    validate_document(questions)

    out = tmp_path / "real_pdf_output.xlsx"
    write_excel(questions, str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]

    by_number = {}
    for row in range(2, ws.max_row + 1):
        by_number[ws.cell(row=row, column=1).value] = row

    row12 = by_number[12]
    explanation_12 = ws.cell(row=row12, column=18).value  # Question Explanation
    assert r"$4 \times 3 + 1 = 13$" in explanation_12
    assert isinstance(explanation_12, str)
    assert not explanation_12.startswith("=")

    row29 = by_number[29]
    stem_29 = ws.cell(row=row29, column=5).value  # Question Text
    assert r"$\div$" in stem_29 and r"$\times$" in stem_29
    assert not stem_29.startswith("=")

    # Every cell across every math-bearing column, whole document: plain
    # text, never a formula, whenever it contains a LaTeX delimiter.
    checked = 0
    for row in range(2, ws.max_row + 1):
        for col in (5, 8, 10, 12, 14, 18):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and "$" in v:
                checked += 1
                assert not v.startswith("=")
    assert checked > 0  # sanity: we actually found LaTeX-bearing cells
