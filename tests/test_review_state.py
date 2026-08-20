"""
test_review_state.py — unit tests for the human review/edit engine
(review_state.py). These are the automated equivalents of the manual
acceptance tests: edit+save, reset, image-question option entry,
answer-key-conflict resolution, multi-paper isolation, and reset-all —
all verified against the REAL parsed SSC/AFCAT data, never a toy fixture,
so they exercise the exact objects the UI would hand to Excel export.
"""

import copy

from excel_writer import write_excel, build_marking_scheme, QUESTION_TYPE_OPTIONS
from paper_segmenter import detect_paper_spans
from parser import parse_paper
from review_state import PaperState, compute_blocking_errors
from validator import validate_document


def _build_paper_state(doc_result, span):
    result = parse_paper(doc_result, span)
    questions = result["questions"]
    validate_document(questions)  # annotates confidence/validation_flags in place
    return PaperState(span, result["mode"], questions, {}, result["paper_anomalies"])


def _ssc_paper_state(ssc_doc_result):
    span = detect_paper_spans(ssc_doc_result)[0]
    return _build_paper_state(ssc_doc_result, span)


def _afcat_paper_states(afcat_doc_result):
    return [_build_paper_state(afcat_doc_result, span) for span in detect_paper_spans(afcat_doc_result)]


# ---------------------------------------------------------------------------
# TEST 1 — normal edit: edit Q1's text, save, verify effective + Excel output
# ---------------------------------------------------------------------------

def test_edit_question_text_reaches_effective_and_excel(ssc_doc_result, tmp_path):
    ps = _ssc_paper_state(ssc_doc_result)
    original_text = ps.get_current(1)["stem"]
    new_text = "EDITED — manually corrected question text for Q1."

    ps.apply_edit(1, {"stem": new_text})

    assert ps.get_current(1)["stem"] == new_text
    assert ps.is_modified(1)
    # original snapshot must never change
    assert ps.original_by_number[1]["stem"] == original_text

    effective = ps.effective_questions()
    q1 = next(q for q in effective if q["number"] == 1)
    assert q1["stem"] == new_text

    out = str(tmp_path / "ssc_edited.xlsx")
    write_excel(effective, out, marking_scheme=build_marking_scheme(len(effective), 200, None))

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert new_text in ws.cell(row=2, column=5).value


# ---------------------------------------------------------------------------
# TEST 2 — reset restores the original value, including in the Excel export
# ---------------------------------------------------------------------------

def test_reset_question_restores_original(ssc_doc_result, tmp_path):
    ps = _ssc_paper_state(ssc_doc_result)
    original_text = ps.get_current(1)["stem"]

    ps.apply_edit(1, {"stem": "temporary edit"})
    assert ps.is_modified(1)

    ps.reset_question(1)
    assert not ps.is_modified(1)
    assert ps.get_current(1)["stem"] == original_text
    assert ps.status_badges(1) in (["✓ Extracted"], ["⚠ Needs Review"])

    out = str(tmp_path / "ssc_reset.xlsx")
    write_excel(ps.effective_questions(), out, marking_scheme=build_marking_scheme(200, 200, None))
    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=5).value.strip().startswith(original_text.split("\n")[0][:20])


# ---------------------------------------------------------------------------
# TEST 3 — marking scheme overrides reach the Excel output
# ---------------------------------------------------------------------------

def test_marking_scheme_override_in_excel(ssc_doc_result, tmp_path):
    ps = _ssc_paper_state(ssc_doc_result)
    effective = ps.effective_questions()

    marking_scheme = build_marking_scheme(len(effective), max_marks=200, negative_marks=None)
    # user manually overrides marks-per-question and negative marks
    marking_scheme["marks_per_question"] = 2.5
    marking_scheme["negative_marks"] = 0.5

    out = str(tmp_path / "ssc_marking.xlsx")
    write_excel(effective, out, marking_scheme=marking_scheme)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=6).value == 2.5
    assert ws.cell(row=2, column=7).value == 0.5


# ---------------------------------------------------------------------------
# TEST 4 — needs-review question (AFCAT Q23 typo) gets corrected and
# flips to "Manually Confirmed" while remembering it was flagged.
# ---------------------------------------------------------------------------

def test_needs_review_question_can_be_corrected(afcat_doc_result):
    ps = _afcat_paper_states(afcat_doc_result)[0]  # Paper 1
    q23 = ps.get_current(23)
    assert ps.was_flagged(23), "Q23 should be originally flagged (incomplete option list)"
    assert ps.status_badges(23) == ["⚠ Needs Review"]

    ps.apply_edit(23, {
        "options": {**q23["options"], "C": "Thorough"},
    })

    assert ps.is_modified(23)
    assert ps.was_flagged(23)  # never forgotten
    assert ps.status_badges(23) == ["✓ Manually Confirmed", "✎ Modified"]
    assert ps.get_current(23)["options"]["C"] == "Thorough"

    counts = ps.review_counts()
    assert counts["resolved"] >= 1
    assert counts["remaining"] == counts["needs_review"] - counts["resolved"]


def test_answer_key_conflict_resolution_via_dropdown():
    """A question flagged with an answer-key conflict must not have its
    correct_answer silently picked — only an explicit edit resolves it."""
    fake_question = {
        "number": 1, "sequence_number": 1, "section": "SECTION-A", "q_type": "MCQ",
        "stem": "Q", "passage": None, "options": {"A": "a", "B": "b", "C": "c", "D": "d"},
        "is_image": False, "anomaly_notes": ["answer key conflict: compact=A, solutions=B"],
        "char_start": 0, "char_end": 1, "source_page": 1, "source_method": "native",
        "source_confidence": 1.0, "correct_answer": "", "explanation": "",
    }
    validate_document([fake_question])
    assert fake_question["confidence"] == "LOW"

    class FakeSpan:
        paper_id = 1
        paper_name = None
        status = "OK"
        ambiguity_reason = None

    ps = PaperState(FakeSpan(), "digit", [fake_question], {}, [])
    assert ps.was_flagged(1)
    assert ps.get_current(1)["correct_answer"] == ""

    ps.apply_edit(1, {"correct_answer": "B"})
    assert ps.get_current(1)["correct_answer"] == "B"
    assert ps.status_badges(1) == ["✓ Manually Confirmed", "✎ Modified"]


# ---------------------------------------------------------------------------
# TEST 5 — image question: no fabricated options, manual entry works
# ---------------------------------------------------------------------------

def test_image_question_manual_option_entry(afcat_doc_result):
    paper_states = _afcat_paper_states(afcat_doc_result)
    ps = paper_states[0]
    image_qs = [n for n in ps.question_order if ps.get_current(n)["is_image"]]
    assert image_qs, "Paper 1 should have at least one image question"
    n = image_qs[0]

    original = ps.get_current(n)
    assert all(v == "" for v in original["options"].values()), "must start blank, never fabricated"

    ps.apply_edit(n, {"options": {"A": "Triangle", "B": "Square", "C": "Circle", "D": "Hexagon"}})
    updated = ps.get_current(n)
    assert updated["options"]["A"] == "Triangle"
    assert ps.is_modified(n)

    effective = ps.effective_questions()
    q = next(q for q in effective if q["number"] == n)
    assert q["options"]["A"] == "Triangle"


# ---------------------------------------------------------------------------
# TEST 6 — multi-paper isolation: editing Paper 2 must never touch Paper 1
# ---------------------------------------------------------------------------

def test_multi_paper_isolation(afcat_doc_result):
    paper_states = _afcat_paper_states(afcat_doc_result)
    p1, p2 = paper_states[0], paper_states[1]

    p1_q23_before = copy.deepcopy(p1.get_current(23))
    p2.apply_edit(23, {"stem": "PAPER 2 EDIT ONLY"})

    assert p2.get_current(23)["stem"] == "PAPER 2 EDIT ONLY"
    assert p1.get_current(23) == p1_q23_before
    assert not p1.is_modified(23)
    assert p2.is_modified(23)


def test_multi_paper_isolation_in_excel_export(afcat_doc_result, tmp_path):
    paper_states = _afcat_paper_states(afcat_doc_result)
    p1, p2 = paper_states[0], paper_states[1]
    p2.apply_edit(23, {"stem": "PAPER 2 ONLY EDIT MARKER"})

    out1 = str(tmp_path / "paper1.xlsx")
    out2 = str(tmp_path / "paper2.xlsx")
    write_excel(p1.effective_questions(), out1, marking_scheme=build_marking_scheme(100, 300, 1))
    write_excel(p2.effective_questions(), out2, marking_scheme=build_marking_scheme(100, 300, 1))

    import openpyxl
    wb1 = openpyxl.load_workbook(out1)["questions"]
    wb2 = openpyxl.load_workbook(out2)["questions"]
    text1 = "\n".join(str(wb1.cell(row=r, column=5).value) for r in range(2, wb1.max_row + 1))
    text2 = "\n".join(str(wb2.cell(row=r, column=5).value) for r in range(2, wb2.max_row + 1))
    assert "PAPER 2 ONLY EDIT MARKER" not in text1
    assert "PAPER 2 ONLY EDIT MARKER" in text2


# ---------------------------------------------------------------------------
# TEST 7 — reset all edits restores every question in that paper
# ---------------------------------------------------------------------------

def test_reset_all_restores_every_question(afcat_doc_result):
    ps = _afcat_paper_states(afcat_doc_result)[0]
    originals = {n: copy.deepcopy(ps.get_current(n)) for n in ps.question_order}

    ps.apply_edit(1, {"stem": "edit 1"})
    ps.apply_edit(2, {"stem": "edit 2"})
    ps.apply_edit(23, {"options": {"A": "x", "B": "y", "C": "z", "D": "w"}})
    assert len(ps.edited) == 3

    ps.reset_all()

    assert len(ps.edited) == 0
    for n in ps.question_order:
        assert ps.get_current(n) == originals[n]


# ---------------------------------------------------------------------------
# Blocking-validation logic
# ---------------------------------------------------------------------------

def test_blocking_errors_detects_duplicate_and_invalid_answer():
    questions = [
        {"number": 1, "correct_answer": "A"},
        {"number": 1, "correct_answer": "Z"},
    ]
    errors = compute_blocking_errors(questions, {"maximum_marks": 100, "marks_per_question": 1, "negative_marks": 0})
    assert any("Duplicate" in e for e in errors)
    assert any("not A/B/C/D" in e for e in errors)


def test_blocking_errors_detects_corrupted_marking_scheme():
    questions = [{"number": 1, "correct_answer": "A"}]
    errors = compute_blocking_errors(questions, {"maximum_marks": -5, "marks_per_question": 0, "negative_marks": -1})
    assert any("Maximum Marks" in e for e in errors)
    assert any("Marks per Question" in e for e in errors)
    assert any("Negative Marks" in e for e in errors)


def test_blocking_errors_clean_data_has_none(afcat_doc_result):
    ps = _afcat_paper_states(afcat_doc_result)[0]
    errors = compute_blocking_errors(ps.effective_questions(), {"maximum_marks": 300, "marks_per_question": 3, "negative_marks": 1})
    assert errors == []


def test_question_type_options_include_existing_template_values():
    assert "objective_single_choice" in QUESTION_TYPE_OPTIONS
    assert len(QUESTION_TYPE_OPTIONS) == 6
