"""
tests/test_question_stem_visual_separation.py — regression tests for the
Q49/Q50 class of problem: a question's STEM can contain a genuine visual
(often a thin-line-drawn grid/table/matrix diagram) while its OPTIONS are
completely ordinary text. Detection of each destination (question, option
A-D) must be fully independent and spatial — never inferred from
q["is_image"], never merged into one screenshot, never propagated from
one destination to another.

Root cause fixed (see image_locator.py's _rect_overlap_fraction and
_is_significant_bbox docstrings): PyMuPDF represents each straight-line
stroke of a hand-drawn grid (e.g. Q50's number grid, built from 36
individual border strokes) as a rect with ZERO width or height. The old
area-ratio overlap formula divided by that zero area and discarded every
such stroke before it could ever be attributed to a region — a real
diagram made entirely of thin strokes produced ZERO candidates, not a
merged one. Fixed by handling degenerate rects explicitly and checking
"is this significant" on the UNION of a band's rects, not on each
individual rect beforehand.
"""

from __future__ import annotations

from pathlib import Path

import fitz
import pytest

from excel_writer import write_excel, TEMPLATE_PATH
from image_extractor import build_artifact
from image_locator import locate_image_regions

ROOT = Path(__file__).resolve().parent.parent
SSC_PDF = str(ROOT / "SSC Stenographer MTP-2.pdf")
MIXED_FIXTURE = str(Path(__file__).resolve().parent / "fixtures" / "synthetic_stem_visual_mixed.pdf")


@pytest.fixture(scope="module")
def ssc_page3():
    doc = fitz.open(SSC_PDF)
    yield doc[2]  # page 3 (1-indexed) — Q49/Q50's page
    doc.close()


@pytest.fixture(scope="module")
def mixed_fixture():
    doc = fitz.open(MIXED_FIXTURE)
    yield doc
    doc.close()


# -- Test 1/7 core acceptance criterion: SSC Q50 (real PDF) -----------------

def test_q50_question_visual_detected_and_tightly_bounded(ssc_page3):
    candidates = locate_image_regions(ssc_page3, 50, 51)
    dests = {c.destination for c in candidates}
    assert dests == {"question"}  # ONLY the question — never merged with options
    q = next(c for c in candidates if c.destination == "question")
    assert q.confidence == "high"
    # the grid's real extent, confirmed by direct PDF inspection — must
    # NOT extend down into the option row (which starts at y~452).
    assert q.bbox[1] >= 415 and q.bbox[3] <= 451


def test_q50_crop_contains_only_the_grid_not_the_options(ssc_page3):
    candidates = locate_image_regions(ssc_page3, 50, 51)
    q = next(c for c in candidates if c.destination == "question")
    artifact = build_artifact(ssc_page3, q)
    assert artifact is not None
    # a tight grid crop (padded) is a small image, not a whole-region
    # screenshot — sanity bound on rendered pixel dimensions at the
    # module's RENDER_DPI (220): the real grid bbox is ~152x28pt, so at
    # 220dpi that's roughly 460x85px — nowhere near a full question+
    # options region (which would be several times taller).
    from PIL import Image
    import io
    img = Image.open(io.BytesIO(artifact.image_bytes))
    assert img.height < 150  # would be 300+ if the option row got included


def test_q50_options_remain_pure_text_via_parser(ssc_page3):
    """The parser's own extraction (untouched by this fix) already reads
    Q50's options as text — this asserts that fact stays true and ties it
    to the acceptance criterion's exact expected values."""
    import sys
    sys.path.insert(0, str(ROOT))
    from extractor import extract_document
    from paper_segmenter import detect_paper_spans
    from parser import parse_paper

    doc_result = extract_document(SSC_PDF)
    spans = detect_paper_spans(doc_result)
    result = parse_paper(doc_result, spans[0])
    q50 = next(q for q in result["questions"] if q["number"] == 50)
    assert q50["is_image"] is False
    assert q50["options"] == {"A": "30", "B": "24", "C": "20", "D": "28"}


# -- Test: SSC Q49 (real PDF) — two-matrix stem visual -----------------------

def test_q49_question_visual_detected_no_option_bleed(ssc_page3):
    candidates = locate_image_regions(ssc_page3, 49, 50)
    dests = {c.destination for c in candidates}
    assert dests == {"question"}
    q = candidates[0]
    assert q.confidence == "high"
    # spans both matrices (the question genuinely needs both to answer),
    # but stays well above the option row (which starts further down).
    assert q.bbox[3] < 390  # Q50's own marker starts at y~390.8


# -- Test: mixed case — stem visual + 2 image options + 2 text options -----

def test_mixed_stem_visual_plus_two_image_two_text_options(mixed_fixture):
    page = mixed_fixture[0]
    candidates = {c.destination: c for c in locate_image_regions(page, 1, 2)}
    assert set(candidates.keys()) == {"question", "option_a", "option_c"}
    assert "option_b" not in candidates  # text-only — never fabricated
    assert "option_d" not in candidates  # text-only — never fabricated
    assert candidates["question"].source_type == "pdf_region_render"
    assert candidates["option_a"].source_type == "embedded_raster"
    assert candidates["option_c"].source_type == "embedded_raster"


# -- Test: stem visual + all 4 image options -> 5 independent destinations --

def test_stem_visual_plus_four_image_options_five_destinations(mixed_fixture):
    page = mixed_fixture[1]
    candidates = {c.destination: c for c in locate_image_regions(page, 1, 2)}
    assert set(candidates.keys()) == {"question", "option_a", "option_b", "option_c", "option_d"}
    boxes = [c.bbox for c in candidates.values()]
    assert len(boxes) == len(set(boxes))  # all five are distinct regions


# -- Test: a text-grid-of-numbers question with NO real drawings ----------

def test_text_only_grid_shaped_question_never_fabricates_a_visual():
    """SSC Q9 ('43, 45, 42, 46, 41, 47, ?' — a number-series question) has
    numbers formatted across multiple lines in its stem but is drawn with
    NO actual vector paths — confirms the thin-line-union fix doesn't
    start hallucinating visuals out of ordinary multi-line text."""
    doc = fitz.open(SSC_PDF)
    try:
        page = doc[0]  # Q9 is on page 1
        candidates = locate_image_regions(page, 9, 10)
        assert candidates == []
    finally:
        doc.close()


# -- Excel verification (acceptance criterion) -------------------------------

def test_excel_q50_shaped_question_question_url_only_options_stay_text(tmp_path):
    q = {
        "number": 50, "sequence_number": 50, "section": "General Intelligence",
        "q_type": "MCQ", "stem": "In the following question, select the number...",
        "passage": None,
        "options": {"A": "30", "B": "24", "C": "20", "D": "28"},
        "is_image": False,  # exactly like the real parser output for Q50
        "anomaly_notes": [], "correct_answer": "A", "explanation": "",
        "images": [{
            "destination": "question", "source": "auto_detected", "region": None,
            "confidence": "high", "image_bytes_hash": "sha-q50", "content_type": "image/png",
            "s3_key": "mock-tests/images/q50.png", "image_url": "https://example.com/q50-grid.png",
            "status": "uploaded", "error": None, "user_confirmed": True,
        }],
    }
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=17).value == "https://example.com/q50-grid.png"  # Question Image URL
    for col in (9, 11, 13, 15):  # Option Image URL A-D
        assert ws.cell(row=2, column=col).value in (None, "")
    assert ws.cell(row=2, column=8).value == "30"   # Option A text
    assert ws.cell(row=2, column=10).value == "24"  # Option B text
    assert ws.cell(row=2, column=12).value == "20"  # Option C text
    assert ws.cell(row=2, column=14).value == "28"  # Option D text
