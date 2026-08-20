"""
test_difficulty.py — Difficulty Level bulk assignment (parsing,
assignment, conflict/invalid detection) and the "required Excel values"
amendment (Negative Marks defaults to 0, image questions never fabricate
option text, Question Type stays a real template value).

Uses the REAL parsed SSC/AFCAT data (via the same fixtures as the other
suites), never a toy fixture, so these exercise the exact objects the UI
hands to Excel export.
"""

import openpyxl

from excel_writer import write_excel, build_marking_scheme, QUESTION_TYPE_OPTIONS, DIFFICULTY_LEVEL_OPTIONS
from paper_segmenter import detect_paper_spans
from parser import parse_paper
from review_state import PaperState, compute_blocking_errors, parse_question_number_list
from validator import validate_document


def _build_paper_state(doc_result, span):
    result = parse_paper(doc_result, span)
    questions = result["questions"]
    validate_document(questions)
    return PaperState(span, result["mode"], questions, {}, result["paper_anomalies"])


def _ssc_paper_state(ssc_doc_result):
    span = detect_paper_spans(ssc_doc_result)[0]
    return _build_paper_state(ssc_doc_result, span)


def _afcat_paper_states(afcat_doc_result):
    return [_build_paper_state(afcat_doc_result, span) for span in detect_paper_spans(afcat_doc_result)]


# 1. Parsing -----------------------------------------------------------

def test_parse_mixed_ranges_and_singles():
    result = parse_question_number_list("1-5, 8, 10-12", valid_numbers=set(range(1, 101)))
    assert result.numbers == {1, 2, 3, 4, 5, 8, 10, 11, 12}
    assert result.errors == []


def test_parse_malformed_entries_produce_clear_errors():
    valid = set(range(1, 101))
    assert parse_question_number_list("abc", valid).errors
    assert parse_question_number_list("5-", valid).errors
    r = parse_question_number_list("10-2", valid)
    assert r.errors and "malformed range" in r.errors[0]
    assert r.numbers == set()  # a malformed range contributes no numbers


# 2. Easy/Medium/Hard assignment ----------------------------------------

def test_easy_medium_hard_map_to_questions(ssc_doc_result):
    ps = _ssc_paper_state(ssc_doc_result)
    ps.set_difficulty_text("easy", "1-50")
    ps.set_difficulty_text("medium", "51-150")
    ps.set_difficulty_text("hard", "151-200")

    summary = ps.difficulty_summary()
    assert summary["counts"] == {"easy": 50, "medium": 100, "hard": 50}
    assert summary["resolved"][1] == "easy"
    assert summary["resolved"][100] == "medium"
    assert summary["resolved"][200] == "hard"
    assert summary["unassigned"] == []
    assert summary["conflicts"] == {}


# 3. Duplicate / conflicting assignment ----------------------------------

def test_conflicting_assignment_is_detected_and_blocking(ssc_doc_result):
    ps = _ssc_paper_state(ssc_doc_result)
    ps.set_difficulty_text("easy", "1-20")
    ps.set_difficulty_text("medium", "15-30")

    summary = ps.difficulty_summary()
    assert set(summary["conflicts"]) == set(range(15, 21))

    errs = ps.difficulty_errors()
    assert any("15" in e and ("Easy" in e or "Medium" in e) for e in errs)

    blocking = compute_blocking_errors(
        ps.effective_questions(), {"maximum_marks": 200, "marks_per_question": 1, "negative_marks": 0},
        difficulty_errors=errs,
    )
    assert blocking  # export must be blocked


# 4. Invalid question numbers ---------------------------------------------

def test_invalid_question_number_is_detected(ssc_doc_result):
    ps = _ssc_paper_state(ssc_doc_result)
    ps.set_difficulty_text("easy", "1-5, 9999")  # 9999 doesn't exist in this 200-question paper
    summary = ps.difficulty_summary()
    assert any("9999" in e for e in summary["results"]["easy"].errors)
    assert 9999 not in summary["resolved"]

    errs = ps.difficulty_errors()
    assert any("9999" in e for e in errs)


# 5. Unassigned difficulty --------------------------------------------------

def test_unassigned_difficulty_is_detected_but_not_blocking(ssc_doc_result):
    ps = _ssc_paper_state(ssc_doc_result)
    ps.set_difficulty_text("easy", "1-190")  # leaves 191-200 unassigned
    summary = ps.difficulty_summary()
    assert summary["unassigned"] == list(range(191, 201))

    # unassigned alone is NOT a blocking error (only conflicts/invalid are)
    assert ps.difficulty_errors() == []
    blocking = compute_blocking_errors(
        ps.effective_questions(), {"maximum_marks": 200, "marks_per_question": 1, "negative_marks": 0},
        difficulty_errors=ps.difficulty_errors(),
    )
    assert blocking == []


# 6/7. Negative Marks default vs preserved --------------------------------

def test_negative_marks_defaults_to_zero_when_not_specified(ssc_doc_result, tmp_path):
    ps = _ssc_paper_state(ssc_doc_result)
    effective = ps.effective_questions()
    ms = build_marking_scheme(len(effective), max_marks=200, negative_marks=None)
    assert ms["negative_marks"] is None  # NOT invented at the marking-scheme layer

    out = str(tmp_path / "neg_default.xlsx")
    write_excel(effective, out, marking_scheme=ms)
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=7).value == 0  # but the exported cell is never blank


def test_negative_marks_preserved_when_specified(afcat_doc_result, tmp_path):
    ps = _afcat_paper_states(afcat_doc_result)[0]
    effective = ps.effective_questions()
    ms = build_marking_scheme(len(effective), max_marks=300, negative_marks=1)

    out = str(tmp_path / "neg_present.xlsx")
    write_excel(effective, out, marking_scheme=ms)
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=7).value == 1


# 8. Edited question + difficulty assignment both reach Excel -------------

def test_edit_and_difficulty_both_reach_excel(ssc_doc_result, tmp_path):
    ps = _ssc_paper_state(ssc_doc_result)
    ps.apply_edit(1, {"stem": "EDITED STEM FOR DIFFICULTY TEST"})
    ps.set_difficulty_text("easy", "1-50")
    ps.set_difficulty_text("medium", "51-150")
    ps.set_difficulty_text("hard", "151-200")

    effective = ps.effective_questions()
    ps.apply_difficulty(effective)

    out = str(tmp_path / "edit_and_difficulty.xlsx")
    write_excel(effective, out, marking_scheme=build_marking_scheme(len(effective), 200, None))
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=5).value == "EDITED STEM FOR DIFFICULTY TEST"
    assert ws.cell(row=2, column=4).value == "easy"
    assert ws.cell(row=200, column=4).value == "hard"


# 9. Multi-paper difficulty isolation --------------------------------------

def test_multi_paper_difficulty_isolation(afcat_doc_result):
    paper_states = _afcat_paper_states(afcat_doc_result)
    p1, p2 = paper_states[0], paper_states[1]

    p1.set_difficulty_text("easy", "1-30")
    p2.set_difficulty_text("hard", "1-30")

    assert p1.difficulty_summary()["resolved"].get(1) == "easy"
    assert p2.difficulty_summary()["resolved"].get(1) == "hard"
    assert p1.difficulty_text != p2.difficulty_text
    assert p2.difficulty_summary()["resolved"].get(1) != "easy"


# 10. Image question: Question Type populated, no fabricated options ------

def test_image_question_type_populated_no_fabricated_options(afcat_doc_result, tmp_path):
    ps = _afcat_paper_states(afcat_doc_result)[0]
    image_q_number = next(n for n in ps.question_order if ps.get_current(n)["is_image"])

    effective = ps.effective_questions()
    ps.apply_difficulty(effective)  # unassigned is fine here — not the point of this test
    q = next(q for q in effective if q["number"] == image_q_number)
    assert all(v == "" for v in q["options"].values())  # never fabricated

    out = str(tmp_path / "image_question.xlsx")
    write_excel(effective, out, marking_scheme=build_marking_scheme(len(effective), 300, 1))
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    row = 1 + image_q_number
    assert ws.cell(row=row, column=3).value in QUESTION_TYPE_OPTIONS
    assert ws.cell(row=row, column=3).value == "objective_single_choice"
    assert ws.cell(row=row, column=5).value.startswith("[IMAGE QUESTION")
    for col in (8, 10, 12, 14):  # Option A/B/C/D columns
        assert ws.cell(row=row, column=col).value in ("", None)


def test_difficulty_level_options_match_template():
    assert DIFFICULTY_LEVEL_OPTIONS == ["easy", "medium", "hard"]
