"""
test_app_ui.py — exercises the ACTUAL app.py Streamlit code (widget keys,
form handling, session-state wiring, rerun logic), not just the
underlying review_state engine, using Streamlit's headless AppTest
framework (no browser available in this environment).

Scope/honesty note: AppTest has no simulated st.file_uploader, so these
tests start from session_state pre-populated exactly as app.py leaves it
right after a real "Analyze PDF" click (same doc_result/PaperState
objects a real upload would produce) — i.e. everything AFTER upload,
which is the entire surface this feature adds, is exercised through the
real rendered app.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest
from streamlit.testing.v1 import AppTest

from extractor import extract_document, detect_marking_scheme_for_range
from paper_segmenter import detect_paper_spans
from parser import parse_paper
from validator import validate_document
from review_state import PaperState

from conftest import SSC_PDF  # single source of truth — templates/sample_input.pdf

ROOT = Path(__file__).resolve().parent.parent
APP_PATH = str(ROOT / "app.py")
AFCAT_PDF = str(ROOT / "templates" / "AFCAT_Mock_1-5_SP_2026_Final_File.pdf")


def _build_paper_states(pdf_path: str):
    doc = extract_document(pdf_path)
    spans = detect_paper_spans(doc)
    states = []
    for span in spans:
        result = parse_paper(doc, span)
        questions = result["questions"]
        validate_document(questions)
        marking = detect_marking_scheme_for_range(pdf_path, span.start_page, span.end_page)
        states.append(PaperState(span, result["mode"], questions, marking, result["paper_anomalies"]))
    return doc, states


def _seeded_app(pdf_path: str, name: str) -> AppTest:
    doc, states = _build_paper_states(pdf_path)
    at = AppTest.from_file(APP_PATH, default_timeout=60)
    at.session_state["analyzed"] = True
    at.session_state["doc_result"] = doc
    at.session_state["paper_states"] = states
    at.session_state["uploaded_name"] = name
    at.session_state["uploaded_key"] = (name, 1)
    at.run()
    assert not at.exception, at.exception
    return at


# ---------------------------------------------------------------------------
# TEST 1 — normal edit, through the real UI: open Q1, edit text, save,
# verify the preview updates, generate Excel, verify the edited text is
# actually in the exported file.
# ---------------------------------------------------------------------------

def test_ui_edit_save_generate_roundtrip(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(SSC_PDF, "SSC.pdf")

    at.button(key="editbtn_1_1").click().run()
    assert not at.exception

    new_text = "UI-EDITED TEXT FOR Q1 — verifies real Streamlit wiring."
    at.text_area[0].set_value(new_text).run()
    at.button(key="FormSubmitter:editform_1_1-Save Changes").click().run()
    assert not at.exception

    # in-memory state updated, preview should now show the edited text
    ps = at.session_state["paper_states"][0]
    assert ps.get_current(1)["stem"] == new_text
    assert ps.is_modified(1)

    gen_buttons = [b for b in at.button if b.key == "gen_btn_1" or b.key == "GENERATE EXCEL ANYWAY"]
    # label may be "GENERATE EXCEL" or "GENERATE EXCEL ANYWAY" depending on
    # remaining flagged questions — key is stable either way.
    at.button(key="gen_btn_1").click().run()
    assert not at.exception
    assert _excel_path(at, 1)

    wb = openpyxl.load_workbook(_excel_path(at, 1))
    ws = wb["questions"]
    assert new_text in ws.cell(row=2, column=5).value


# ---------------------------------------------------------------------------
# TEST 2 — reset through the real UI restores the original text
# ---------------------------------------------------------------------------

def test_ui_reset_question_restores_original(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(SSC_PDF, "SSC.pdf")
    ps = at.session_state["paper_states"][0]
    original_text = ps.original_by_number[1]["stem"]

    at.button(key="editbtn_1_1").click().run()
    at.text_area[0].set_value("temporary UI edit").run()
    at.button(key="FormSubmitter:editform_1_1-Save Changes").click().run()
    assert ps.is_modified(1)

    at.button(key="resetq_1_1").click().run()
    assert not at.exception
    assert not ps.is_modified(1)
    assert ps.get_current(1)["stem"] == original_text

    at.button(key="gen_btn_1").click().run()
    wb = openpyxl.load_workbook(_excel_path(at, 1))
    ws = wb["questions"]
    assert ws.cell(row=2, column=5).value.split("\n")[0] == original_text.split("\n")[0]


# ---------------------------------------------------------------------------
# TEST 6 — multi-paper isolation through the real UI: edit Paper 2's Q23
# via the paper selector, verify Paper 1's Q23 is untouched, and only
# Paper 2's exported Excel contains the edit.
# ---------------------------------------------------------------------------

def test_ui_multi_paper_isolation(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(AFCAT_PDF, "AFCAT.pdf")

    # switch to Paper 2 via the paper selector
    at.selectbox(key="active_paper_idx").select(1).run()
    assert not at.exception

    states = at.session_state["paper_states"]
    p1, p2 = states[0], states[1]
    p1_q23_before = p1.get_current(23)["stem"]

    # navigate to Q23 within the now-active Paper 2 via the question
    # selector (the same widget the search box filters), then open the
    # edit form.
    at.selectbox(key="qselect_2").select(23).run()
    assert at.session_state["qselect_2"] == 23

    at.button(key="editbtn_2_23").click().run()
    assert at.session_state["edit_mode_2"] is True

    at.text_area[0].set_value("PAPER 2 ONLY UI EDIT").run()
    at.button(key="FormSubmitter:editform_2_23-Save Changes").click().run()
    assert not at.exception

    assert p2.get_current(23)["stem"] == "PAPER 2 ONLY UI EDIT"
    assert p1.get_current(23)["stem"] == p1_q23_before  # untouched

    at.button(key="gen_btn_2").click().run()
    assert not at.exception
    out2 = _excel_path(at, 2)
    wb2 = openpyxl.load_workbook(out2)["questions"]
    text2 = "\n".join(str(wb2.cell(row=r, column=5).value) for r in range(2, wb2.max_row + 1))
    assert "PAPER 2 ONLY UI EDIT" in text2

    # switch back to Paper 1 and generate — must NOT contain Paper 2's edit
    at.selectbox(key="active_paper_idx").select(0).run()
    at.button(key="gen_btn_1").click().run()
    assert not at.exception
    out1 = _excel_path(at, 1)
    wb1 = openpyxl.load_workbook(out1)["questions"]
    text1 = "\n".join(str(wb1.cell(row=r, column=5).value) for r in range(2, wb1.max_row + 1))
    assert "PAPER 2 ONLY UI EDIT" not in text1


# ---------------------------------------------------------------------------
# TEST 4 — a genuinely flagged question (AFCAT Paper 1 Q23's real content
# typo) goes through Review/Edit -> correct it -> Save -> status flips to
# Manually Confirmed (while remembering it was originally flagged).
# ---------------------------------------------------------------------------

def test_ui_needs_review_to_manually_confirmed(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(AFCAT_PDF, "AFCAT.pdf")
    ps = at.session_state["paper_states"][0]
    assert ps.was_flagged(23)
    assert ps.status_badges(23) == ["⚠ Needs Review"]

    at.button(key="needsreview_jump_1_23").click().run()
    assert not at.exception
    assert at.session_state["qselect_1"] == 23
    assert at.session_state["edit_mode_1"] is True

    # form field order is Option A, B, C, D, Section (indices 1-5 of
    # text_input; index 0 is the search box) — fill in the missing C.
    at.text_input[3].set_value("Thorough").run()
    at.button(key="FormSubmitter:editform_1_23-Save Changes").click().run()
    assert not at.exception

    assert ps.get_current(23)["options"]["C"] == "Thorough"
    assert ps.status_badges(23) == ["✓ Manually Confirmed", "✎ Modified"]

    at.button(key="validate_btn_1").click().run()
    assert not at.exception
    assert st_key_exists(at, "last_validation_1")


def st_key_exists(at, key):
    try:
        at.session_state[key]
        return True
    except Exception:
        return False


def _excel_path(at, pid):
    """Reads the generated Excel path from the new Fix-2 generation state
    dict (st.session_state[f"generation_{pid}"]) -- replaces the old
    excel_path_{pid} key."""
    state = at.session_state[f"generation_{pid}"]
    assert state["status"] == "success", state
    return state["path"]


# ---------------------------------------------------------------------------
# TEST 5 — image question: options start blank (never fabricated), user
# manually enters one, and it reaches the Excel export.
# ---------------------------------------------------------------------------

def test_ui_image_question_manual_option_entry(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(AFCAT_PDF, "AFCAT.pdf")
    ps = at.session_state["paper_states"][0]
    image_q = next(n for n in ps.question_order if ps.get_current(n)["is_image"])
    assert all(v == "" for v in ps.get_current(image_q)["options"].values())

    at.selectbox(key="qselect_1").select(image_q).run()
    assert "🖼" in at.markdown[0].value or any("IMAGE QUESTION" in m.value for m in at.markdown)

    at.button(key=f"editbtn_1_{image_q}").click().run()
    assert not at.exception
    at.text_input[1].set_value("Triangle").run()  # Option A
    at.button(key=f"FormSubmitter:editform_1_{image_q}-Save Changes").click().run()
    assert not at.exception

    assert ps.get_current(image_q)["options"]["A"] == "Triangle"

    at.button(key="gen_btn_1").click().run()
    wb = openpyxl.load_workbook(_excel_path(at, 1))
    ws = wb["questions"]
    row = 1 + image_q
    assert ws.cell(row=row, column=8).value == "Triangle"  # Option A column


# ---------------------------------------------------------------------------
# TEST 3 — marking scheme edits reach the exported Excel
# ---------------------------------------------------------------------------

def test_ui_marking_scheme_edit_reaches_excel(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(SSC_PDF, "SSC.pdf")

    at.text_input(key="mpq_1").set_value("2.5").run()
    at.text_input(key="neg_1").set_value("0.5").run()
    assert not at.exception

    at.button(key="gen_btn_1").click().run()
    assert not at.exception
    wb = openpyxl.load_workbook(_excel_path(at, 1))
    ws = wb["questions"]
    assert ws.cell(row=2, column=6).value == 2.5
    assert ws.cell(row=2, column=7).value == 0.5


# ---------------------------------------------------------------------------
# TEST 7 — Reset All Edits (with confirmation) restores every question
# ---------------------------------------------------------------------------

def test_ui_reset_all_edits_with_confirmation(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    at = _seeded_app(SSC_PDF, "SSC.pdf")
    ps = at.session_state["paper_states"][0]

    # edit Q1 through the real Save flow
    at.button(key="editbtn_1_1").click().run()
    at.text_area[0].set_value("edit one").run()
    at.button(key="FormSubmitter:editform_1_1-Save Changes").click().run()
    assert not at.exception

    # edit Q5 through the real Save flow
    at.selectbox(key="qselect_1").select(5).run()
    at.button(key="editbtn_1_5").click().run()
    at.text_area[0].set_value("edit two").run()
    at.button(key="FormSubmitter:editform_1_5-Save Changes").click().run()
    assert not at.exception
    assert len(ps.edited) == 2

    at.button(key="resetall_btn_1").click().run()
    assert not at.exception
    assert st_key_exists(at, "confirm_reset_all_1")

    at.button(key="resetall_confirm_1").click().run()
    assert not at.exception
    assert len(ps.edited) == 0


# ---------------------------------------------------------------------------
# TEST 8 (V3) — image question: full auto-detect -> confirm -> (mocked) S3
# upload -> URL reaches Excel, through the REAL rendered UI (button clicks,
# not direct PaperState calls). AWS is entirely mocked — no real network
# call, no real credentials, per the V3 testing plan.
# ---------------------------------------------------------------------------

def test_ui_image_question_auto_detect_confirm_uploads_and_reaches_excel(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "fake-key")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "fake-secret")
    monkeypatch.setenv("AWS_REGION", "ap-south-1")
    monkeypatch.setenv("AWS_S3_BUCKET", "ai-education-s3-public-media")
    monkeypatch.setenv("AWS_S3_KEY_PREFIX", "mock-tests/images/")

    import aws_uploader
    from botocore.exceptions import ClientError

    class FakeS3Client:
        def __init__(self):
            self.objects = {}
            self.put_calls = []

        def head_object(self, Bucket, Key):
            if Key not in self.objects:
                raise ClientError({"Error": {"Code": "404", "Message": "Not Found"}}, "HeadObject")
            return {}

        def put_object(self, Bucket, Key, Body, ContentType):
            self.put_calls.append((Bucket, Key, ContentType))
            self.objects[Key] = Body

    fake_client = FakeS3Client()
    monkeypatch.setattr(aws_uploader, "build_client", lambda config: fake_client)

    doc, states = _build_paper_states(SSC_PDF)
    at = AppTest.from_file(APP_PATH, default_timeout=60)
    at.session_state["analyzed"] = True
    at.session_state["doc_result"] = doc
    at.session_state["paper_states"] = states
    at.session_state["uploaded_name"] = "SSC.pdf"
    at.session_state["uploaded_key"] = ("SSC.pdf", 1)
    at.session_state["pdf_bytes"] = Path(SSC_PDF).read_bytes()  # V3: needed for image detection
    at.run()
    assert not at.exception

    ps = states[0]
    assert ps.get_current(5)["is_image"]
    assert ps.get_images(5) == []  # nothing confirmed yet

    at.selectbox(key="qselect_1").select(5).run()
    assert not at.exception

    # candidate index 0 is the "question" destination (dict insertion order
    # from image_locator's bands: question, option_a, option_b, option_c,
    # option_d) — confirmed by image_locator's own tests.
    assert at.button(key="usedet_1_5_0")
    at.button(key="usedet_1_5_0").click().run()
    assert not at.exception

    images = ps.get_images(5)
    question_images = [im for im in images if im["destination"] == "question"]
    assert len(question_images) == 1
    assert question_images[0]["status"] == "uploaded"
    url = question_images[0]["image_url"]
    assert url.startswith(
        "https://ai-education-s3-public-media.s3.ap-south-1.amazonaws.com/mock-tests/images/"
    )
    assert len(fake_client.put_calls) == 1  # exactly one real "upload"

    # re-rendering the same question again must NOT re-upload (dedup cache)
    at.selectbox(key="qselect_1").select(6).run()
    at.selectbox(key="qselect_1").select(5).run()
    assert not at.exception
    assert len(fake_client.put_calls) == 1

    at.button(key="gen_btn_1").click().run()
    assert not at.exception
    wb = openpyxl.load_workbook(_excel_path(at, 1))
    ws = wb["questions"]
    assert ws.cell(row=1 + 5, column=17).value == url  # Question Image URL column


# ---------------------------------------------------------------------------
# TEST 9 (V3) — ZIP export preserves each paper's own confirmed image URL,
# with no cross-paper bleed, through the real "GENERATE ALL (ZIP)" button.
# ---------------------------------------------------------------------------

def test_ui_zip_export_preserves_per_paper_image_urls(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "fake-key")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "fake-secret")
    monkeypatch.setenv("AWS_REGION", "ap-south-1")
    monkeypatch.setenv("AWS_S3_BUCKET", "ai-education-s3-public-media")
    monkeypatch.setenv("AWS_S3_KEY_PREFIX", "mock-tests/images/")

    import aws_uploader
    from botocore.exceptions import ClientError

    class FakeS3Client:
        def __init__(self):
            self.objects = {}

        def head_object(self, Bucket, Key):
            if Key not in self.objects:
                raise ClientError({"Error": {"Code": "404", "Message": "Not Found"}}, "HeadObject")
            return {}

        def put_object(self, Bucket, Key, Body, ContentType):
            self.objects[Key] = Body

    monkeypatch.setattr(aws_uploader, "build_client", lambda config: FakeS3Client())

    doc, states = _build_paper_states(AFCAT_PDF)
    at = AppTest.from_file(APP_PATH, default_timeout=60)
    at.session_state["analyzed"] = True
    at.session_state["doc_result"] = doc
    at.session_state["paper_states"] = states
    at.session_state["uploaded_name"] = "AFCAT.pdf"
    at.session_state["uploaded_key"] = ("AFCAT.pdf", 1)
    at.session_state["pdf_bytes"] = Path(AFCAT_PDF).read_bytes()
    at.run()
    assert not at.exception

    p1, p2 = states[0], states[1]
    p1_image_q = next(n for n in p1.question_order if p1.get_current(n)["is_image"])
    p2_image_q = next(n for n in p2.question_order if p2.get_current(n)["is_image"])

    # confirm an image for paper 1's image question
    at.selectbox(key="qselect_1").select(p1_image_q).run()
    at.button(key=f"usedet_1_{p1_image_q}_0").click().run()
    assert not at.exception
    url1 = next(im["image_url"] for im in p1.get_images(p1_image_q) if im["status"] == "uploaded")

    # switch to paper 2 and confirm an image for ITS image question
    at.selectbox(key="active_paper_idx").select(1).run()
    assert not at.exception
    at.selectbox(key="qselect_2").select(p2_image_q).run()
    at.button(key=f"usedet_2_{p2_image_q}_0").click().run()
    assert not at.exception
    url2 = next(im["image_url"] for im in p2.get_images(p2_image_q) if im["status"] == "uploaded")

    assert url1 != url2  # distinct diagrams -> distinct content hashes -> distinct URLs

    zip_button = next(b for b in at.button if b.label == "GENERATE ALL (ZIP)")
    zip_button.click().run()
    assert not at.exception
    zip_state = at.session_state["zip_generation"]
    assert zip_state["status"] == "success"
    zip_bytes = zip_state["bytes"]
    assert zip_bytes

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        paper1_name = next(n for n in names if "Paper1" in n)
        paper2_name = next(n for n in names if "Paper2" in n)

        wb1 = openpyxl.load_workbook(io.BytesIO(zf.read(paper1_name)))["questions"]
        wb2 = openpyxl.load_workbook(io.BytesIO(zf.read(paper2_name)))["questions"]

    assert wb1.cell(row=1 + p1_image_q, column=17).value == url1
    assert wb2.cell(row=1 + p2_image_q, column=17).value == url2
    # no cross-paper bleed: paper 1's file must not carry paper 2's URL, and vice versa
    all_urls_wb1 = [wb1.cell(row=r, column=17).value for r in range(2, wb1.max_row + 1)]
    all_urls_wb2 = [wb2.cell(row=r, column=17).value for r in range(2, wb2.max_row + 1)]
    assert url2 not in all_urls_wb1
    assert url1 not in all_urls_wb2


# ---------------------------------------------------------------------------
# TEST 10 — SSC Q50 through the real UI: a question with is_image=False
# (options extracted as real text) but a genuine stem-embedded diagram.
# Before this fix, the image section never rendered at all for such a
# question (gated solely on is_image). Confirms: the section now appears
# based on spatial detection alone, confirming it uploads ONLY to
# Question Image URL, and the option TEXT cells are completely untouched.
# ---------------------------------------------------------------------------

def test_ui_q50_stem_visual_detected_without_is_image_flag(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "templates").symlink_to(ROOT / "templates")
    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "fake-key")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "fake-secret")
    monkeypatch.setenv("AWS_REGION", "ap-south-1")
    monkeypatch.setenv("AWS_S3_BUCKET", "ai-education-s3-public-media")
    monkeypatch.setenv("AWS_S3_KEY_PREFIX", "mock-tests/images/")

    import aws_uploader
    from botocore.exceptions import ClientError

    class FakeS3Client:
        def __init__(self):
            self.objects = {}
            self.put_calls = []

        def head_object(self, Bucket, Key):
            if Key not in self.objects:
                raise ClientError({"Error": {"Code": "404", "Message": "Not Found"}}, "HeadObject")
            return {}

        def put_object(self, Bucket, Key, Body, ContentType):
            self.put_calls.append((Bucket, Key, ContentType))
            self.objects[Key] = Body

    fake_client = FakeS3Client()
    monkeypatch.setattr(aws_uploader, "build_client", lambda config: fake_client)

    doc, states = _build_paper_states(SSC_PDF)
    at = AppTest.from_file(APP_PATH, default_timeout=60)
    at.session_state["analyzed"] = True
    at.session_state["doc_result"] = doc
    at.session_state["paper_states"] = states
    at.session_state["uploaded_name"] = "SSC.pdf"
    at.session_state["uploaded_key"] = ("SSC.pdf", 1)
    at.session_state["pdf_bytes"] = Path(SSC_PDF).read_bytes()
    at.run()
    assert not at.exception

    ps = states[0]
    assert ps.get_current(50)["is_image"] is False  # the parser signal that must NOT gate this
    assert ps.get_current(50)["options"] == {"A": "30", "B": "24", "C": "20", "D": "28"}

    at.selectbox(key="qselect_1").select(50).run()
    assert not at.exception
    # the new spatial-detection label, not the legacy "IMAGE QUESTION" one
    assert any("Visual Content Detected" in m.value for m in at.markdown)

    assert at.button(key="usedet_1_50_0")
    at.button(key="usedet_1_50_0").click().run()
    assert not at.exception

    images = ps.get_images(50)
    assert len(images) == 1
    assert images[0]["destination"] == "question"
    assert images[0]["status"] == "uploaded"
    url = images[0]["image_url"]
    assert url.startswith("https://ai-education-s3-public-media.s3.ap-south-1.amazonaws.com/mock-tests/images/")
    assert len(fake_client.put_calls) == 1

    at.button(key="gen_btn_1").click().run()
    assert not at.exception
    state = at.session_state["generation_1"]
    assert state["status"] == "success"

    wb = openpyxl.load_workbook(state["path"])
    ws = wb["questions"]
    row = 1 + 50
    assert ws.cell(row=row, column=17).value == url  # Question Image URL
    for col in (9, 11, 13, 15):  # Option Image URL A-D — never populated
        assert ws.cell(row=row, column=col).value in (None, "")
    assert ws.cell(row=row, column=8).value == "30"   # Option A text, untouched
    assert ws.cell(row=row, column=10).value == "24"  # Option B text, untouched
    assert ws.cell(row=row, column=12).value == "20"  # Option C text, untouched
    assert ws.cell(row=row, column=14).value == "28"  # Option D text, untouched

