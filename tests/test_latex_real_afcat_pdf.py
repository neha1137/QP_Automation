"""
tests/test_latex_real_afcat_pdf.py — regression tests anchored to the exact
real-world bug report: the cylindrical-water-tank question in
templates/AFCAT_Mock_1-5_SP_2026_Final_File.pdf (Sample Question Paper-1,
Q39), whose explanation is authored as an embedded equation-editor object.

Before the mathematical-content cleanup/reconstruction stage
(math_reconstructor.py) existed, this PDF's native text extraction alone
produced fragmented numbers ("3 5" instead of "3.5"), a 22/7 fraction split
across two separate lines with no "/" character at all, and literal
Private-Use-Area tofu-box glyphs (the font's stretchy-parenthesis pieces)
— which latex_processor.py then had no way to safely turn into correct
LaTeX, per the bug report this feature fixes.

These tests exercise the FULL real pipeline: extractor.py (now including
math_reconstructor.py) -> parser.py -> latex_processor.py -> excel_writer.py,
against the actual PDF bytes — not a synthetic string.
"""

from __future__ import annotations

import re

import openpyxl

import parser as parser_mod
from excel_writer import write_excel, TEMPLATE_PATH
from paper_segmenter import detect_paper_spans
from validator import validate_document

_PUA_RE = re.compile(r"[-]")


def _paper1_questions(afcat_doc_result):
    spans = detect_paper_spans(afcat_doc_result)
    result = parser_mod.parse_paper(afcat_doc_result, spans[0])
    return {q["number"]: q for q in result["questions"]}


def _tank_question(afcat_doc_result):
    q = _paper1_questions(afcat_doc_result)[39]
    assert "cylindrical water tank" in q["stem"].lower()
    return q


# -- The real tank question / explanation --------------------------------

def test_tank_question_stem_preserves_pi_and_fraction(afcat_doc_result):
    q = _tank_question(afcat_doc_result)
    assert r"\pi = \frac{22}{7}" in q["stem"]
    assert "3.5" in q["stem"]
    assert "8 m" in q["stem"]


def test_tank_question_explanation_has_correct_formula(afcat_doc_result):
    q = _tank_question(afcat_doc_result)
    explanation = q["explanation"]
    # Volume of a cylinder = πr²h, correctly represented as LaTeX.
    assert r"$\pi r^2h$" in explanation
    # The Given-line fraction plus both derivation-line fractions.
    assert explanation.count(r"\frac{22}{7}") >= 3
    # The decimal values survive intact — never fragmented ("3 5"/"12 25").
    assert "3.5" in explanation
    assert "12.25" in explanation
    assert "3 5" not in explanation
    assert "12 25" not in explanation
    # The final numeric results are preserved exactly.
    assert "= 308 cubic metres." in explanation
    assert r"$308 \times 1000 = 308,000$" in explanation


def test_tank_question_explanation_contains_the_exact_reconstructed_equations(afcat_doc_result):
    """The core requirement of the equation-block reconstruction stage:
    the full "Volume = ..." derivation — previously fragmented across a
    dozen scrambled lines with tofu-box glyphs — must now read as exactly
    these three equations, matching the source PDF's own mathematical
    meaning (Volume of a cylinder = πr²h)."""
    q = _tank_question(afcat_doc_result)
    explanation = q["explanation"]
    assert r"Volume = $\frac{22}{7} \times (3.5)^2 \times 8$" in explanation
    assert r"= $\frac{22}{7} \times 12.25 \times 8$" in explanation
    assert "= 308 cubic metres." in explanation


def test_tank_question_explanation_has_no_stray_standalone_numbers(afcat_doc_result):
    """3.5, 8, 12.25 and the exponent 2 must appear ONLY inside the
    reconstructed equation — never left behind as their own orphaned
    line/token outside it (the exact fragmentation this feature fixes:
    'Volume = $\\frac{22}{7}$\\n3.5\\n8\\n...\\n2\\n(\\n)$\\times$(...')."""
    q = _tank_question(afcat_doc_result)
    explanation = q["explanation"]
    lines = [ln.strip() for ln in explanation.split("\n")]
    for orphan in ("3.5", "8", "12.25", "2"):
        assert orphan not in lines, f"{orphan!r} left as a standalone line: {lines}"


def test_tank_question_explanation_has_no_empty_parentheses(afcat_doc_result):
    q = _tank_question(afcat_doc_result)
    assert "()" not in q["explanation"]
    assert not re.search(r"\(\s*\)", q["explanation"])


def test_tank_question_explanation_has_no_pdf_font_artifacts(afcat_doc_result):
    """The stretchy-parenthesis tofu-box glyphs from the equation-editor
    font must never reach Streamlit/Excel as raw Private-Use-Area
    characters — they are either reconstructed into literal parentheses
    or (if unconfident) left as genuinely visible source text, never as
    invisible/garbled codepoints."""
    q = _tank_question(afcat_doc_result)
    assert not _PUA_RE.search(q["stem"])
    assert not _PUA_RE.search(q["explanation"])
    for letter in "ABCD":
        assert not _PUA_RE.search(q["options"][letter])


def test_tank_question_explanation_never_splits_a_number_with_a_dollar_sign(afcat_doc_result):
    q = _tank_question(afcat_doc_result)
    assert "$/" not in q["explanation"]
    assert not re.search(r"\$\d[\d,]*,\d+(?!\$)", q["explanation"])


# -- Whole-document invariants for this PDF, not just Q39 ------------------

def test_no_pua_artifacts_anywhere_in_afcat_paper1(afcat_doc_result):
    questions = _paper1_questions(afcat_doc_result)
    offenders = []
    for q in questions.values():
        fields = [("stem", q.get("stem")), ("explanation", q.get("explanation"))]
        fields += [(f"option_{L}", q["options"].get(L)) for L in "ABCD"]
        for name, text in fields:
            if text and _PUA_RE.search(text):
                offenders.append((q["number"], name))
    assert offenders == []


def test_no_english_word_ever_appears_inside_a_latex_span_afcat_paper1(afcat_doc_result):
    latex_span_re = re.compile(r"\$(.+?)\$", re.S)
    english_word_re = re.compile(r"[A-Za-z]{3,}")
    safe_words = {"times", "div", "le", "ge", "neq", "pm", "sqrt", "frac", "pi"}
    questions = _paper1_questions(afcat_doc_result)
    offenders = []
    for q in questions.values():
        fields = [("stem", q.get("stem")), ("explanation", q.get("explanation"))]
        fields += [(f"option_{L}", q["options"].get(L)) for L in "ABCD"]
        for name, text in fields:
            if not text:
                continue
            for m in latex_span_re.finditer(text):
                bad = [w for w in english_word_re.findall(m.group(1)) if w.lower() not in safe_words]
                if bad:
                    offenders.append((q["number"], name, m.group(0), bad))
    assert offenders == []


def test_image_detection_unaffected_by_math_reconstruction(afcat_doc_result):
    """Sanity anchor for requirement 14: the image pipeline (is_image
    classification) must be completely independent of the LaTeX/math-
    reconstruction feature, even on this new document."""
    questions = _paper1_questions(afcat_doc_result)
    image_qs = {n for n, q in questions.items() if q["is_image"]}
    # Ground truth confirmed by hand against the real PDF's own images.
    assert image_qs == {52, 59, 60, 72, 73}


# -- Excel: real PDF -> parser -> LaTeX -> Excel cell, end to end ---------

def test_tank_question_excel_cell_is_plain_text_latex(afcat_doc_result, tmp_path):
    questions = list(_paper1_questions(afcat_doc_result).values())
    validate_document(questions)

    out = tmp_path / "afcat_paper1_output.xlsx"
    write_excel(questions, str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    by_number = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}

    row39 = by_number[39]
    stem_cell = ws.cell(row=row39, column=5)
    explanation_cell = ws.cell(row=row39, column=18)

    assert isinstance(stem_cell.value, str)
    assert isinstance(explanation_cell.value, str)
    assert stem_cell.data_type == "s"
    assert explanation_cell.data_type == "s"
    assert not stem_cell.value.startswith("=")
    assert not explanation_cell.value.startswith("=")
    assert r"\frac{22}{7}" in stem_cell.value
    assert r"\pi r^2h" in explanation_cell.value
    assert "= 308 cubic metres." in explanation_cell.value
    assert not explanation_cell.value.startswith("=")
    assert explanation_cell.data_type != "f"
    assert r"Volume = $\frac{22}{7} \times (3.5)^2 \times 8$" in explanation_cell.value
    assert r"= $\frac{22}{7} \times 12.25 \times 8$" in explanation_cell.value
    assert not _PUA_RE.search(explanation_cell.value)
    assert "()" not in explanation_cell.value
