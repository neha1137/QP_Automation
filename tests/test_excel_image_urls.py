"""
tests/test_excel_image_urls.py — excel_writer.py: confirmed image URLs
land in the correct existing template columns, unconfirmed/ambiguous ones
never do, and multi-paper isolation holds for image state exactly as it
already does for every other field.

Column numbers below match excel_writer.HEADER_ORDER exactly (never
re-invented here): 9=Option Image URL A, 11=B, 13=C, 15=D, 17=Question
Image URL.
"""

from __future__ import annotations

import openpyxl
import pytest

from excel_writer import write_excel, TEMPLATE_PATH


def _base_question(number: int, images: list[dict] | None = None) -> dict:
    return {
        "number": number,
        "sequence_number": number,
        "section": "General",
        "q_type": "IMAGE_MCQ",
        "stem": f"Question {number} stem",
        "passage": None,
        "options": {"A": "", "B": "", "C": "", "D": ""},
        "is_image": True,
        "anomaly_notes": [],
        "correct_answer": "A",
        "explanation": "",
        "images": images or [],
    }


def _uploaded(destination: str, url: str) -> dict:
    return {
        "destination": destination, "source": "auto_detected", "region": None,
        "confidence": "high", "image_bytes_hash": "sha", "content_type": "image/png",
        "s3_key": f"mock-tests/images/{destination}.png", "image_url": url,
        "status": "uploaded", "error": None, "user_confirmed": True,
    }


def _failed(destination: str) -> dict:
    return {
        "destination": destination, "source": "auto_detected", "region": None,
        "confidence": "high", "image_bytes_hash": "sha", "content_type": "image/png",
        "s3_key": None, "image_url": None,
        "status": "failed", "error": "boom", "user_confirmed": False,
    }


def test_question_image_url_lands_in_correct_column(tmp_path):
    q = _base_question(1, images=[_uploaded("question", "https://example.com/q1.png")])
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=17).value == "https://example.com/q1.png"  # Question Image URL
    for col in (9, 11, 13, 15):  # option image columns must stay blank
        assert ws.cell(row=2, column=col).value in (None, "")


@pytest.mark.parametrize("destination,column", [
    ("option_a", 9), ("option_b", 11), ("option_c", 13), ("option_d", 15),
])
def test_each_option_image_url_lands_in_its_own_column(tmp_path, destination, column):
    url = f"https://example.com/{destination}.png"
    q = _base_question(1, images=[_uploaded(destination, url)])
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=column).value == url
    assert ws.cell(row=2, column=17).value in (None, "")  # question image column stays blank


def test_question_and_option_images_coexist_without_cross_contamination(tmp_path):
    q = _base_question(1, images=[
        _uploaded("question", "https://example.com/q.png"),
        _uploaded("option_b", "https://example.com/b.png"),
    ])
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=17).value == "https://example.com/q.png"
    assert ws.cell(row=2, column=11).value == "https://example.com/b.png"
    assert ws.cell(row=2, column=9).value in (None, "")
    assert ws.cell(row=2, column=13).value in (None, "")
    assert ws.cell(row=2, column=15).value in (None, "")


def test_unconfirmed_or_failed_image_never_reaches_excel(tmp_path):
    q = _base_question(1, images=[_failed("question")])
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=17).value in (None, "")  # never a fabricated/failed URL


def test_no_images_key_at_all_writes_blank_unchanged_from_v2(tmp_path):
    """A plain V2-shaped question dict (no 'images' key at all) must still
    write successfully with blank image columns — backward compatible."""
    q = _base_question(1, images=None)
    del q["images"]
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    for col in (9, 11, 13, 15, 17):
        assert ws.cell(row=2, column=col).value in (None, "")


def test_text_only_question_never_gets_image_columns_populated(tmp_path):
    q = _base_question(1)
    q["is_image"] = False
    q["options"] = {"A": "Alpha", "B": "Beta", "C": "Gamma", "D": "Delta"}
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    for col in (9, 11, 13, 15, 17):
        assert ws.cell(row=2, column=col).value in (None, "")


def test_multi_question_image_isolation(tmp_path):
    """Two different questions' image state never bleeds into each
    other's row — mirrors the existing per-paper isolation guarantee."""
    q1 = _base_question(1, images=[_uploaded("question", "https://example.com/1.png")])
    q2 = _base_question(2, images=[_uploaded("question", "https://example.com/2.png")])
    out = tmp_path / "out.xlsx"
    write_excel([q1, q2], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=17).value == "https://example.com/1.png"
    assert ws.cell(row=3, column=17).value == "https://example.com/2.png"
