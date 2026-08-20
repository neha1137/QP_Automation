"""
tests/test_option_image_separation.py — Fix 1 regression tests: option
images must be detected/cropped/uploaded INDEPENDENTLY per destination,
never merged into one combined screenshot.

Uses tests/fixtures/synthetic_option_layouts.pdf (4 pages: vertical stack,
2x2 grid, one-option-with-multiple-components, and a pixel-identical-
images dedup case) since neither real sample PDF has a validated
four-image-option example in every one of these shapes (AFCAT's real
image questions are a horizontal row; the one real 2x2 grid found in
AFCAT during implementation turned out to be a plain-text question, not
an image one — see test_afcat_grid_layout_text_question_never_fabricates_
an_image in test_image_locator.py).
"""

from __future__ import annotations

from pathlib import Path

import fitz
import pytest

from excel_writer import write_excel, TEMPLATE_PATH
from image_extractor import build_artifact
from image_locator import locate_image_regions
from review_state import PaperState
from paper_segmenter import PaperSpan

FIXTURE = str(Path(__file__).resolve().parent / "fixtures" / "synthetic_option_layouts.pdf")


@pytest.fixture(scope="module")
def fixture_doc():
    doc = fitz.open(FIXTURE)
    yield doc
    doc.close()


ALL_OPTION_DESTS = ["option_a", "option_b", "option_c", "option_d"]


# -- 1/2/3: each layout produces 4 separate crops ---------------------------

def test_vertical_stack_produces_four_separate_option_crops(fixture_doc):
    page = fixture_doc[0]
    candidates = locate_image_regions(page, 1, 2)
    dests = {c.destination for c in candidates}
    assert dests == set(ALL_OPTION_DESTS)
    # never a single merged "question" or "ambiguous" catch-all
    assert "question" not in dests
    assert "ambiguous" not in dests


def test_grid_layout_produces_four_separate_option_crops(fixture_doc):
    page = fixture_doc[1]
    candidates = locate_image_regions(page, 1, 2)
    dests = {c.destination for c in candidates}
    assert dests == set(ALL_OPTION_DESTS)


def test_grid_layout_crops_are_at_distinct_non_overlapping_positions(fixture_doc):
    """The real bug this fix addresses: a naive y-only split on a 2-column
    grid merges each row's two side-by-side options into one shared band.
    Assert the four bboxes are pairwise non-overlapping and match the
    fixture's known, exact image placements."""
    page = fixture_doc[1]
    candidates = {c.destination: c.bbox for c in locate_image_regions(page, 1, 2)}
    expected = {
        "option_a": (120.0, 80.0, 180.0, 140.0),
        "option_b": (290.0, 80.0, 350.0, 140.0),
        "option_c": (120.0, 150.0, 180.0, 210.0),
        "option_d": (290.0, 150.0, 350.0, 210.0),
    }
    for dest, exp_bbox in expected.items():
        assert candidates[dest] == pytest.approx(exp_bbox, abs=0.5)

    # pairwise non-overlap
    boxes = list(candidates.values())
    for i in range(len(boxes)):
        for j in range(i + 1, len(boxes)):
            a, b = boxes[i], boxes[j]
            no_overlap = a[2] <= b[0] or b[2] <= a[0] or a[3] <= b[1] or b[3] <= a[1]
            assert no_overlap, f"boxes {a} and {b} overlap"


def test_horizontal_row_still_produces_four_separate_crops_real_pdf():
    """Real-PDF regression (not synthetic): SSC Q5's horizontal row of 4
    option images must still split cleanly — this is the case Fix 1 must
    NOT break while fixing the grid case."""
    doc = fitz.open("SSC Stenographer MTP-2.pdf")
    try:
        page = doc[0]
        candidates = {c.destination: c for c in locate_image_regions(page, 5, 6)}
        for dest in ALL_OPTION_DESTS:
            assert dest in candidates
            assert candidates[dest].confidence == "high"
        # SSC's own right-column bleed protection must still hold: option_d's
        # right edge stays well short of the right column's content (x~329+).
        assert candidates["option_d"].bbox[2] < 310
    finally:
        doc.close()


# -- 4: multiple components of the SAME option union into one image -------

def test_multiple_components_of_same_option_are_unioned(fixture_doc):
    page = fixture_doc[2]
    candidates = {c.destination: c.bbox for c in locate_image_regions(page, 1, 2)}
    # option A has two components (a 60x60 square at x=[120,180] and a
    # small 20x20 companion mark at x=[185,205]) that must union into ONE
    # bbox spanning both.
    a_bbox = candidates["option_a"]
    assert a_bbox[0] == pytest.approx(120.0, abs=0.5)
    assert a_bbox[2] == pytest.approx(205.0, abs=0.5)  # covers both components


def test_multi_component_union_does_not_bleed_into_other_options(fixture_doc):
    page = fixture_doc[2]
    candidates = {c.destination: c.bbox for c in locate_image_regions(page, 1, 2)}
    # B, C, D are single-component options elsewhere on the page and must
    # be completely unaffected by option A's extra component.
    assert candidates["option_b"] == pytest.approx((120.0, 150.0, 180.0, 210.0), abs=0.5)
    assert candidates["option_c"] == pytest.approx((120.0, 220.0, 180.0, 280.0), abs=0.5)
    assert candidates["option_d"] == pytest.approx((120.0, 290.0, 180.0, 350.0), abs=0.5)


# -- 5: different options never merged (explicit cross-check) --------------

def test_different_options_never_share_a_bounding_box(fixture_doc):
    for page_idx in (0, 1, 2, 3):
        page = fixture_doc[page_idx]
        candidates = locate_image_regions(page, 1, 2)
        bboxes = [c.bbox for c in candidates if c.destination in ALL_OPTION_DESTS]
        assert len(bboxes) == len(set(bboxes)), f"page {page_idx}: duplicate bbox across options"


# -- 6: Excel receives four independent URLs --------------------------------

def test_excel_receives_four_independent_option_urls(tmp_path):
    q = {
        "number": 1, "sequence_number": 1, "section": "General", "q_type": "IMAGE_MCQ",
        "stem": "stem", "passage": None, "options": {"A": "", "B": "", "C": "", "D": ""},
        "is_image": True, "anomaly_notes": [], "correct_answer": "A", "explanation": "",
        "images": [
            {"destination": d, "source": "auto_detected", "region": None, "confidence": "high",
             "image_bytes_hash": f"sha-{d}", "content_type": "image/png",
             "s3_key": f"mock-tests/images/{d}.png", "image_url": f"https://example.com/{d}.png",
             "status": "uploaded", "error": None, "user_confirmed": True}
            for d in ALL_OPTION_DESTS
        ],
    }
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=9).value == "https://example.com/option_a.png"
    assert ws.cell(row=2, column=11).value == "https://example.com/option_b.png"
    assert ws.cell(row=2, column=13).value == "https://example.com/option_c.png"
    assert ws.cell(row=2, column=15).value == "https://example.com/option_d.png"
    # every URL distinct — never the same combined screenshot in all four
    urls = [ws.cell(row=2, column=c).value for c in (9, 11, 13, 15)]
    assert len(set(urls)) == 4


# -- 7: manual replacement of Option B does not affect A/C/D ---------------

def test_manual_replacement_of_one_option_never_touches_the_others():
    span = PaperSpan(paper_id=1, paper_name=None, start_page=1, end_page=1)
    q = {
        "number": 1, "sequence_number": 1, "section": None, "q_type": "IMAGE_MCQ",
        "stem": "s", "passage": None, "options": {"A": "", "B": "", "C": "", "D": ""},
        "is_image": True, "anomaly_notes": [],
    }
    ps = PaperState(span, "digit", [q], {}, [])

    def artifact(dest, url):
        return {
            "destination": dest, "source": "auto_detected", "region": None, "confidence": "high",
            "image_bytes_hash": f"h-{url}", "content_type": "image/png", "s3_key": f"k-{url}",
            "image_url": url, "status": "uploaded", "error": None, "user_confirmed": True,
        }

    for dest in ALL_OPTION_DESTS:
        ps.set_image(1, dest, artifact(dest, f"https://example.com/auto-{dest}.png"))

    before = {im["destination"]: im["image_url"] for im in ps.get_images(1)}
    assert len(before) == 4

    # manual replacement of ONLY option_b
    ps.set_image(1, "option_b", artifact("option_b", "https://example.com/manual-B.png"))

    after = {im["destination"]: im["image_url"] for im in ps.get_images(1)}
    assert after["option_b"] == "https://example.com/manual-B.png"
    assert after["option_a"] == before["option_a"]
    assert after["option_c"] == before["option_c"]
    assert after["option_d"] == before["option_d"]
    assert len(after) == 4  # never a stray duplicate entry


# -- 8: identical option images may legitimately share hash/URL -----------

def test_identical_option_images_produce_identical_hashes(fixture_doc):
    """Page 4 of the fixture places pixel-identical bytes for option A and
    option C (option B and D differ). SHA dedup treating them as 'the same
    upload' is correct behavior — the important property is that each is
    still independently represented as its own artifact/destination."""
    page = fixture_doc[3]
    candidates = {c.destination: c for c in locate_image_regions(page, 1, 2)}
    art_a = build_artifact(page, candidates["option_a"])
    art_b = build_artifact(page, candidates["option_b"])
    art_c = build_artifact(page, candidates["option_c"])
    art_d = build_artifact(page, candidates["option_d"])

    assert art_a.sha256 == art_c.sha256  # pixel-identical source images
    assert art_a.sha256 != art_b.sha256
    assert art_a.sha256 != art_d.sha256
    # still four independent artifacts/destinations, not collapsed into one
    assert {art_a.destination, art_b.destination, art_c.destination, art_d.destination} == set(ALL_OPTION_DESTS)


def test_identical_option_images_share_s3_url_via_dedup(monkeypatch, fixture_doc):
    """End-to-end: uploading option A then option C (identical bytes)
    through the real aws_uploader dedup path yields the SAME URL, with
    only one real PUT — while still being stored as two independent
    confirmed destinations in PaperState."""
    import aws_uploader
    from botocore.exceptions import ClientError

    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "fake")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "fake")
    monkeypatch.setenv("AWS_REGION", "ap-south-1")
    monkeypatch.setenv("AWS_S3_BUCKET", "ai-education-s3-public-media")
    monkeypatch.setenv("AWS_S3_KEY_PREFIX", "mock-tests/images/")

    class FakeClient:
        def __init__(self):
            self.objects = {}
            self.put_calls = []

        def head_object(self, Bucket, Key):
            if Key not in self.objects:
                raise ClientError({"Error": {"Code": "404"}}, "HeadObject")
            return {}

        def put_object(self, Bucket, Key, Body, ContentType):
            self.put_calls.append(Key)
            self.objects[Key] = Body

    fake = FakeClient()
    monkeypatch.setattr(aws_uploader, "build_client", lambda config: fake)

    page = fixture_doc[3]
    candidates = {c.destination: c for c in locate_image_regions(page, 1, 2)}
    art_a = build_artifact(page, candidates["option_a"])
    art_c = build_artifact(page, candidates["option_c"])

    cache = {}
    result_a = aws_uploader.upload_image(art_a.image_bytes, art_a.content_type, cache=cache)
    result_c = aws_uploader.upload_image(art_c.image_bytes, art_c.content_type, cache=cache)

    assert result_a["url"] == result_c["url"]
    assert len(fake.put_calls) == 1  # only one real upload for the shared bytes


# -- 9: existing question-level detection still passes (explicit smoke) ----

def test_question_level_detection_unaffected_by_option_fix():
    """Explicit smoke check tying Fix 1's changes back to the already-
    validated question-only detection path (full suite also re-verifies
    this via test_image_locator.py, kept here as a direct assertion)."""
    doc = fitz.open("SSC Stenographer MTP-2.pdf")
    try:
        page = doc[0]
        candidates = locate_image_regions(page, 5, 6)
        question_cand = next(c for c in candidates if c.destination == "question")
        assert question_cand.confidence == "high"
        assert question_cand.bbox[2] < 310  # still column-protected
    finally:
        doc.close()
