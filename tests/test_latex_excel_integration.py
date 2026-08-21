"""
tests/test_latex_excel_integration.py — confirms LaTeX-processed text
lands as plain TEXT in the exact same Excel cells that already hold
Question Text/Option A-D/Explanation, and that it coexists correctly with
the image-URL columns (never replaces them, never gets rendered/embedded
as an image) — the CASE 3 requirement from the feature spec: a question
with both extractable math text AND a genuine visual.
"""

from __future__ import annotations

import openpyxl

from excel_writer import write_excel, TEMPLATE_PATH
from latex_processor import process_question_fields


def _base_question(**overrides) -> dict:
    q = {
        "number": 1, "sequence_number": 1, "section": "General",
        "q_type": "MCQ", "stem": "If x² + y² = 25, find x.", "passage": None,
        "options": {"A": "2", "B": "4", "C": "√16", "D": "16/2"},
        "is_image": False, "anomaly_notes": [], "correct_answer": "A",
        "explanation": "Using a² + b² = c², we solve for x.",
    }
    q.update(overrides)
    return q


def test_excel_cells_contain_latex_text_not_formulas(tmp_path):
    q = _base_question()
    process_question_fields(q)
    out = tmp_path / "out.xlsx"
    write_excel([q], str(out), template_path=TEMPLATE_PATH)

    wb = openpyxl.load_workbook(out)
    ws = wb["questions"]
    assert ws.cell(row=2, column=5).value == "If $x^2 + y^2 = 25$, find x."  # Question Text
    assert ws.cell(row=2, column=8).value == "2"                             # Option A
    assert ws.cell(row=2, column=10).value == "4"                            # Option B
    assert ws.cell(row=2, column=12).value == r"$\sqrt{16}$"                 # Option C
    assert ws.cell(row=2, column=14).value == r"$\frac{16}{2}$"              # Option D
    assert ws.cell(row=2, column=18).value == "Using $a^2 + b^2 = c^2$, we solve for x."  # Explanation

    # Cell values must be plain strings (text), never openpyxl formula
    # objects/ArrayFormula — Excel must store LaTeX as text, not compute it.
    for col in (5, 8, 10, 12, 14, 18):
        val = ws.cell(row=2, column=col).value
        assert isinstance(val, str)
        assert not val.startswith("=")


def test_question_with_math_text_and_confirmed_image_preserves_both():
    """CASE 3 from the spec: extractable math text AND a visual coexist —
    Question Text carries the LaTeX, Question Image URL carries the
    existing image pipeline's URL, independently."""
    q = _base_question(
        stem="If x² + y² = 25, find the required value.",
        images=[{
            "destination": "question", "source": "auto_detected", "region": None,
            "confidence": "high", "image_bytes_hash": "sha-1", "content_type": "image/png",
            "s3_key": "mock-tests/images/q1.png", "image_url": "https://example.com/q1-diagram.png",
            "status": "uploaded", "error": None, "user_confirmed": True,
        }],
    )
    process_question_fields(q)

    import tempfile
    with tempfile.TemporaryDirectory() as d:
        out = f"{d}/out.xlsx"
        write_excel([q], out, template_path=TEMPLATE_PATH)
        wb = openpyxl.load_workbook(out)
        ws = wb["questions"]
        assert ws.cell(row=2, column=5).value == "If $x^2 + y^2 = 25$, find the required value."
        assert ws.cell(row=2, column=17).value == "https://example.com/q1-diagram.png"  # Question Image URL
