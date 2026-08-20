"""
excel_writer.py — Populates the EXACT provided Excel template with parsed
questions. Does not create a new workbook or new columns; loads the
template, writes into the same header order, saves to output/.
"""

from __future__ import annotations

import re

import openpyxl

# XML 1.0 (and therefore .xlsx) disallows most C0 control characters.
# PDF/OCR extraction occasionally leaves stray ones (e.g. bullet glyphs
# that decode to \x03) — strip them rather than let openpyxl raise.
_ILLEGAL_XML_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]"
)


def _clean(value):
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub("", value)
    return value

TEMPLATE_PATH = "templates/question_template.xlsx"
SHEET_NAME = "questions"

# Excel Question Type dropdown only allows these values (confirmed from
# the template's own data validation on column C — never invented). None
# of them is "image", so image questions default to objective_single_choice
# too, with the skip clearly marked in the text — see build_question_text().
QUESTION_TYPE_OPTIONS = [
    "objective_single_choice",
    "objective_multiple_choice",
    "subjective_short",
    "subjective_long",
    "assertion",
    "reasoning",
]
EXCEL_QUESTION_TYPE = QUESTION_TYPE_OPTIONS[0]

# Difficulty Level dropdown likewise only allows these three values
# (confirmed from the template's own data validation on column D — never
# invented, and always lowercase to match exactly). A question with no
# confirmed difficulty assignment writes blank here rather than an
# invented level — see review_state.py's difficulty assignment logic for
# how "unassigned" is surfaced to the user before that can happen.
DIFFICULTY_LEVEL_OPTIONS = ["easy", "medium", "hard"]

HEADER_ORDER = [
    "Question No", "Section", "Question Type", "Difficulty Level",
    "Question Text", "Marks", "Negative Marks",
    "Option A", "Option Image URL A",
    "Option B", "Option Image URL B",
    "Option C", "Option Image URL C",
    "Option D", "Option Image URL D",
    "Correct Answer", "Question Image URL",
    "Question Explanation", "Question Explanation Image",
]


def build_marking_scheme(
    total_questions: int,
    max_marks: int | None = None,
    negative_marks: float | None = None,
) -> dict:
    """
    Derives a marking scheme dict from values actually present in the
    source document — never invents marks/negative-marking from exam-name
    conventions or external knowledge. marks_per_question is only filled
    in when max_marks is known (max_marks / total_questions); everything
    else stays None (-> blank in Excel) unless the PDF actually states it.

    This is intentionally NOT baked into the parser: it's a small, optional
    input to the Excel writer so a future UI can compute/let the user
    confirm this per document instead of it being fixed pipeline logic.
    """
    marks_per_question = None
    marks_source = "not_specified"
    if max_marks is not None and total_questions > 0:
        marks_per_question = max_marks / total_questions
        if marks_per_question == int(marks_per_question):
            marks_per_question = int(marks_per_question)
        marks_source = "derived"

    return {
        "maximum_marks": max_marks,
        "marks_per_question": marks_per_question,
        "negative_marks": negative_marks,
        "source": {
            "maximum_marks": "pdf" if max_marks is not None else "not_specified",
            "marks_per_question": marks_source,
            "negative_marks": "pdf" if negative_marks is not None else "not_specified",
        },
    }


# destination key -> which of the template's 5 image-URL columns it feeds.
# Keys match review_state.PaperState.set_image()'s `destination` values
# exactly (V3) — never invented independently here.
_IMAGE_DESTINATIONS = ("question", "option_a", "option_b", "option_c", "option_d")


def _image_url_for(q: dict, destination: str) -> str:
    """The confirmed (status == 'uploaded') S3 URL for this destination, or
    '' if none has been confirmed yet — NEVER a candidate that the user
    hasn't explicitly accepted/uploaded (an unconfirmed auto-detection
    never reaches q['images'] at all — see review_state.PaperState).
    An 'ambiguous' destination (needs-review, not yet resolved by the
    user into a real destination) never matches any column here either."""
    for im in q.get("images", []) or []:
        if im.get("destination") == destination and im.get("status") == "uploaded" and im.get("image_url"):
            return im["image_url"]
    return ""


def build_question_text(q: dict) -> str:
    """Passage (if any) repeated in full ahead of the question's own stem,
    per spec — no separate passage row is created."""
    parts = []
    if q.get("passage"):
        parts.append(q["passage"])
        parts.append("")  # blank line separator
    if q["is_image"]:
        parts.append(
            "[IMAGE QUESTION — IMAGE REQUIRED: this question's content/options "
            "are graphical and were not extracted. Review the source PDF manually.]"
        )
        parts.append("")
    parts.append(q["stem"])
    return "\n".join(parts).strip()


def write_excel(
    questions: list[dict],
    output_path: str,
    template_path: str = TEMPLATE_PATH,
    marking_scheme: dict | None = None,
) -> str:
    """
    marking_scheme, if given, follows build_marking_scheme()'s shape:
    {"marks_per_question": ..., "negative_marks": ..., ...}. Applied
    uniformly to every row. When omitted (default), Marks/Negative Marks
    stay blank — unchanged from prior behavior.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]

    # sanity check the template header hasn't drifted from what we expect
    actual_header = [ws.cell(row=1, column=c).value for c in range(1, len(HEADER_ORDER) + 1)]
    if actual_header != HEADER_ORDER:
        raise ValueError(
            f"Template header mismatch — refusing to write. "
            f"Expected {HEADER_ORDER}, got {actual_header}"
        )

    marks_value = (marking_scheme or {}).get("marks_per_question")
    negative_marks_value = (marking_scheme or {}).get("negative_marks")
    # Negative marking has a principled, non-fabricated default: "not
    # mentioned in the PDF" and "zero negative marking" are the same real
    # fact for this field, unlike Marks (which stays genuinely unknown/
    # blank if never confirmed) — so this is the one field allowed a
    # default, and it's always exactly 0, never a guessed non-zero value.
    if negative_marks_value is None:
        negative_marks_value = 0

    row = 2
    for q in sorted(questions, key=lambda x: x["sequence_number"]):
        opts = q["options"]
        ws.cell(row=row, column=1, value=_clean(q["number"]))                       # Question No
        ws.cell(row=row, column=2, value=_clean(q.get("section") or ""))            # Section
        # A per-question override (set via the review/edit UI) is honored
        # when present and valid; otherwise the same default as before.
        # This is the only way an edited Question Type reaches the
        # export — everything else about this function is unchanged.
        excel_q_type = q.get("excel_question_type")
        if excel_q_type not in QUESTION_TYPE_OPTIONS:
            excel_q_type = EXCEL_QUESTION_TYPE
        ws.cell(row=row, column=3, value=_clean(excel_q_type))                      # Question Type
        # Difficulty Level: only ever one of the template's own three
        # values (set via the bulk difficulty-assignment UI) or blank for
        # a question the user hasn't assigned yet — never invented here.
        difficulty = q.get("difficulty_level") or ""
        if difficulty and difficulty not in DIFFICULTY_LEVEL_OPTIONS:
            difficulty = ""
        ws.cell(row=row, column=4, value=_clean(difficulty))                        # Difficulty Level
        ws.cell(row=row, column=5, value=_clean(build_question_text(q)))            # Question Text
        ws.cell(row=row, column=6, value=_clean(marks_value if marks_value is not None else ""))            # Marks
        ws.cell(row=row, column=7, value=_clean(negative_marks_value))              # Negative Marks (never blank — see above)
        ws.cell(row=row, column=8, value=_clean(opts.get("A", "")))                 # Option A
        ws.cell(row=row, column=9, value=_clean(_image_url_for(q, "option_a")))     # Option Image URL A
        ws.cell(row=row, column=10, value=_clean(opts.get("B", "")))                # Option B
        ws.cell(row=row, column=11, value=_clean(_image_url_for(q, "option_b")))    # Option Image URL B
        ws.cell(row=row, column=12, value=_clean(opts.get("C", "")))                # Option C
        ws.cell(row=row, column=13, value=_clean(_image_url_for(q, "option_c")))    # Option Image URL C
        ws.cell(row=row, column=14, value=_clean(opts.get("D", "")))                # Option D
        ws.cell(row=row, column=15, value=_clean(_image_url_for(q, "option_d")))    # Option Image URL D
        ws.cell(row=row, column=16, value=_clean(q.get("correct_answer") or ""))    # Correct Answer
        ws.cell(row=row, column=17, value=_clean(_image_url_for(q, "question")))    # Question Image URL
        ws.cell(row=row, column=18, value=_clean(q.get("explanation") or ""))       # Question Explanation
        ws.cell(row=row, column=19, value=_clean(""))                               # Question Explanation Image (V3 scope: question/option images only)
        row += 1

    wb.save(output_path)
    return output_path
